import datetime
import os
import time
import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from openpyxl import worksheet
import requests
from bs4 import BeautifulSoup

print('\033[30m\033[47mДомашнее задание по теме "Обзор сторонних библиотек Python".\033[0m')
print('\033[30m\033[47mЦель: познакомиться с использованием сторонних библиотек в Python и '
      'применить их в различных задачах.\033[0m')
print('\033[30m\033[47mСтудент Крылов Эдуард Васильевич\033[0m')
print('\033[30m\033[47mНачало работы над заданием: 21.09.2024г.\033[0m')
thanks = ('\033[30m\033[47mОкончание работы над заданием: 25.09.2024г.\033[0m\n\033[30m\033[47mБлагодарю '
          'за внимание :-)\033[0m')
print()
"""
Получает текущий курс доллара с https://www.banki.ru/products/currency/usd/?ysclid=m1ce84yqf5733524099
"""


class Current:
    """Парсинг сайта"""

    def __init__(self):
        super().__init__()
        self.ts = 0.1
        self.book = openpyxl.Workbook()
        self.check_list = []
        self.ch_kurses = None
        self.ch_img = None
        self.ch_data = None
        self.ch_name = None
        self.name_sheet: str = "Report"
        self.name_book: str = "Report.xlsx"
        self.list_xlsx = ['A', 'B', 'C', 'D', 'E']

    URL = 'https://www.banki.ru/products/currency/usd/'

    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                            'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36'}

    response = requests.get(URL, headers=header)

    if response.status_code == 200:
        print(f'\033[32mСканируем сайт на предмет курса доллара в "Банках России", статус сайта:'
              f' {response.status_code}\033[0m')

        def check_class(self, ch_name, ch_kurses, ch_data):
            self.ch_name = ch_name
            self.ch_kurses = ch_kurses
            self.ch_data = ch_data
            html = requests.get(self.URL)
            html = html.text
            soup = BeautifulSoup(html, 'html.parser')
            name_bank1 = soup.find_all("div", {"class": ch_name})
            checker = len(name_bank1)
            _1 = 0
            _2 = 1

            for _ in range(checker):
                name_bank = soup.find_all("div", {"class": ch_name})
                name_bank = name_bank[_].text
                data_time = soup.findAll("div", {"class": ch_data})
                data_time = data_time[_].text
                result_purchase = soup.findAll("div", {"class": ch_kurses})
                result_purchase = result_purchase[_1].text
                result_purchase = result_purchase.replace(",", ".")
                result_purchase = result_purchase.replace("₽", "")
                result_purchase = str(result_purchase)
                result_sale = soup.findAll("div", {"class": ch_kurses})
                result_sale = result_sale[_2].text
                result_sale = result_sale.replace(",", ".")
                result_sale = result_sale.replace("₽", "")
                result_sale = str(result_sale)
                self.check_list.append([str(_ + 1), str(name_bank), str(result_purchase),
                                        str(result_sale), str(data_time)])
                _1 += 2
                _2 += 2
                print(f'{_ + 1}. Курс банка \033[34m"{name_bank}": \033[32mпокупка\033[0m - \033[31m{result_purchase}'
                      f'\033[0m, \033[34mпродажа\033[0m - \033[31m{result_sale}\033[0m, {data_time}')
                time.sleep(self.ts)
            self.new_data()
    else:
        print(f'\033[32mНе удалось установить соединение с сайтом, код: {response.status_code}\033[0m')

    def new_data(self):
        print(f'\033[32mЗаписываем данные из полученных результатов в "{self.name_book}"\033[0m')
        # os.chdir('./Docs')
        book = openpyxl.load_workbook("Report.xlsx")
        sheet: worksheet = book.worksheets[0]
        s = 3
        line = 0
        for i in self.check_list:
            col = 1
            sheet.cell(row=s, column=col, value=i[line])
            print(f'Ячейка: {self.list_xlsx[line] + str(s)}: {i[line]}')
            col += 1
            time.sleep(self.ts)
            sheet.cell(row=s, column=col, value=i[line + 1])
            print(f'Ячейка: {self.list_xlsx[line + 1] + str(s)}: {i[line + 1]}')
            col += 1
            time.sleep(self.ts)
            if i[line + 2] == '-':
                i[line + 2] = '0'
            sheet.cell(row=s, column=col, value=i[line + 2])
            print(f'Ячейка: {self.list_xlsx[line + 2] + str(s)}: {i[line + 2]}')
            col += 1
            time.sleep(self.ts)
            if i[line + 3] == '-':
                i[line + 3] = '0'
            sheet.cell(row=s, column=col, value=i[line + 3])
            print(f'Ячейка: {self.list_xlsx[line + 3] + str(s)}: {i[line + 3]}')
            col += 1
            time.sleep(self.ts)
            sheet.cell(row=s, column=col, value=i[line + 4])
            print(f'Ячейка: {self.list_xlsx[line + 4] + str(s)}: {i[line + 4]}')
            time.sleep(self.ts)
            s += 1
        book.save(self.name_book)


class Documents(Current):
    """ Создание документа и запись в него полученных данных"""

    def __init__(self):
        super().__init__()
        self.name4 = None
        self.name3 = None
        self.name2 = None
        self.name1 = None
        self.name5 = None
        self.dt_now = datetime.datetime.now()
        self.book = openpyxl.Workbook()

    def doc_save(self, name1, name2, name3, name4, name5):
        self.name1 = name1
        self.name2 = name2
        self.name3 = name3
        self.name4 = name4
        self.name5 = name5
        if os.path.exists('Docs'):
            os.chdir('./Docs')
            print('\033[32mОткрываем папку "Docs"\033[0m')
        else:
            print('\033[31mСоздаем папку "Docs"\033[0m')
            os.mkdir('Docs')
            os.chdir('Docs')

        if os.path.exists(self.name_book):
            print(f'\033[32mОткрываем файл: "{self.name_book}"\033[0m')
            self.new_sheet()
        else:
            print(f'\033[31mСоздаем файл: "{self.name_book}"\033[0m')
            self.book.save(self.name_book)
            self.new_sheet()

    def new_sheet(self):
        print(f'Записываем заголовки в "{self.name_book}"')
        book = openpyxl.load_workbook(filename=self.name_book)
        sheet: worksheet = book.worksheets[0]
        print('Имя листа: ', self.book.sheetnames)
        sheet['A1'].value = 'Дата:'
        sheet['B1'].value = self.dt_now
        sheet['A2'].value = '№ П/П'
        sheet['B2'].value = self.name1
        sheet['C2'].value = self.name2
        sheet['D2'].value = self.name3
        sheet['E2'].value = self.name4
        sheet['F2'].value = self.name5
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["E"].width = 30
        sheet.column_dimensions["F"].width = 30
        book.save(self.name_book)


class OpenDoc(Current):
    """ Считываем данные из текстового документа"""

    def __init__(self):
        super().__init__()
        self.g_banks = []
        self.g_purchase = []
        self.g_sale = []

        print(f'\033[32mСчитываем данные из "{self.name_book}", вычисляем доход банка и записуваем в файл\033[0m')
        book = openpyxl.load_workbook(filename=self.name_book)
        sheet = book.worksheets[0]
        for a in range(3, 13):
            name = sheet['B' + str(a)].value
            self.g_banks.append(name)
            purchase = sheet['C' + str(a)].value
            purchase = float(purchase)
            self.g_purchase.append(purchase)
            sale = sheet['D' + str(a)].value
            sale = float(sale)
            self.g_sale.append(sale)
            income = round(sale - purchase, 2)
            income = str(income)
            sheet['F' + str(a)].value = income
            print(f'Доход банка: \033[34m"{name}"\033[0m при покупке: \033[32m{purchase}\033[0m '
                  f'и продаже: \033[32m{sale}\033[0m - составляет: \033[31m{income}\033[0m')
            time.sleep(self.ts)

        print(f'\033[32mСтроим график из полученных данных\033[0m')
        book.save(self.name_book)
        book.close()

        s_purchase = np.array(self.g_purchase)
        size_purchase = s_purchase.size
        sum_purchase = s_purchase.sum()
        average_purchase = sum_purchase / size_purchase
        print('Среднее значение продаж', average_purchase)

        s_sale = np.array(self.g_sale)
        size_sale = s_sale.size
        sum_sale = s_sale.sum()
        average_sale = sum_sale / size_sale
        print('Среднее значение покупки', average_sale)

        x = self.g_banks
        y1 = self.g_purchase
        y2 = self.g_sale
        y3 = round(average_purchase, 2)
        y4 = round(average_sale, 2)

        plt.title(f'Курсы $ в банках.\nСредняя цена:\nПокупка: {y3} / Продажа: {y4}.')
        plt.xlabel('Наименование банка')
        plt.ylabel('Покупка / Продажа')
        plt.grid()
        plt.plot(x, y1, x, y2, label='Продажа', marker='o')
        plt.show()
        os.startfile(self.name_book)


document = Documents()

document.doc_save('Наименование банка', 'Покупка', 'Продажа',
                  'Дата обновления', 'Прибыль банка')

cur = Current()
cur.check_class("gfTHqP", "jzaqdw", "hDxmZl")

read = OpenDoc()

print()
print(thanks)
